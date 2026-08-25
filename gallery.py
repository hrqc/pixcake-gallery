# -*- coding: utf-8 -*-
"""像素蛋糕选片系统 - 主服务器.
HTTP 层使用标准库，图片处理依赖 Pillow、OpenCV 和 NumPy.

用法:
  python gallery.py [--port 8888] [--ws <工作区根>] [--token <令牌>] [--no-auth]

花生壳远程访问: 花生壳映射本机端口 -> 公网域名/端口, 打开打印的项目链接即可.
"""
import sys, io, os, re, json, time, math, threading, argparse, secrets, hmac, hashlib, urllib.request
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs, quote
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace', line_buffering=True)

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import db
import scanner
import license as lic
from image_service import ImageService, ImageServiceError
from jobs import ExportManager

DATA_DIR = os.path.join(HERE, 'data')
CLEAN_CACHE_DIR = os.path.join(DATA_DIR, 'clean-cache')
QR_DIR = os.path.join(DATA_DIR, 'qr')
TOKEN_FILE = os.path.join(DATA_DIR, 'token.txt')
PLATFORM_FILE = os.path.join(DATA_DIR, 'platform.json')   # 管理端: 平台服务器连接配置
INDEX_HTML = os.path.join(HERE, 'static', 'index.html')
PLATFORM_HTML = os.path.join(HERE, 'static', 'platform.html')   # 管理端平台控制台
DELIVERY_HTML = os.path.join(HERE, 'static', 'delivery.html')
ACTIVATE_HTML = os.path.join(HERE, 'static', 'activate.html')
TENANT_HTML = os.path.join(HERE, 'static', 'tenant.html')
SECRET_FILE = os.path.join(DATA_DIR, 'server_secret')
TENANT_ROOT = os.path.join(DATA_DIR, 'tenants')   # 每个摄影师租户一个子目录
ADMIN_OWNER = 'admin'                             # 管理员自身相册的 owner 标记

PREWARM_HTML = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>预热状态 · 像素蛋糕</title>
<style>
  body{font-family:-apple-system,'Microsoft YaHei',sans-serif;background:#0f1419;color:#e6edf3;margin:0;padding:24px}
  h1{font-size:20px;margin:0 0 18px}
  .row{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:20px}
  .card{background:#161b22;border:1px solid #30363d;border-radius:12px;padding:16px 20px;min-width:130px}
  .card .v{font-size:26px;font-weight:700;margin-top:6px}
  .card .k{font-size:12px;color:#8b949e}
  .card.bad .v{color:#f85149}.card.ok .v{color:#3fb950}.card.run .v{color:#d29922}
  #st{display:inline-block;padding:4px 12px;border-radius:20px;font-size:14px;margin-bottom:18px}
  #st.run{background:#3fb95022;color:#3fb950;border:1px solid #3fb95055}
  #st.idle{background:#8b949e22;color:#8b949e;border:1px solid #8b949e55}
  #st.err{background:#f8514922;color:#f85149;border:1px solid #f8514955}
  ul.fail{list-style:none;padding:0;margin:0}
  ul.fail li{background:#161b22;border:1px solid #30363d;border-radius:8px;padding:8px 12px;margin-bottom:8px;font-size:13px}
  ul.fail code{color:#f85149}
  .muted{color:#8b949e;font-size:12px}
  a{color:#58a6ff;text-decoration:none}
</style>
</head>
<body>
<h1>自动预热 · 去水印图生成状态</h1>
<div id="st" class="run">加载中…</div>
<div class="row">
  <div class="card"><div class="k">相册照片总数</div><div class="v" id="cTotal">–</div></div>
  <div class="card ok"><div class="k">去水印已完成</div><div class="v" id="cDone">–</div></div>
  <div class="card"><div class="k">本轮新构建</div><div class="v" id="cBuilt">–</div></div>
  <div class="card bad"><div class="k">失败</div><div class="v" id="cFailed">–</div></div>
</div>
<div class="muted">上次扫描：<span id="cScan">–</span>　|　当前阶段：<span id="cPhase">–</span></div>
<h2 style="font-size:16px;margin:24px 0 10px">失败照片（<span id="fCount">0</span>）</h2>
<ul class="fail" id="fList"></ul>
<div class="muted" style="margin-top:20px"><a href="/">← 返回管理端</a>　自动每 3 秒刷新</div>
<script>
async function load(){
  try{
    const j = await fetch('/api/prewarm?t=__TOKEN__').then(r=>r.json());
    const st = document.getElementById('st');
    st.className = j.phase==='error' ? 'err' : (j.running ? 'run' : 'idle');
    st.textContent = j.phase==='error' ? '异常' : (j.running ? '运行中' : '空闲');
    document.getElementById('cTotal').textContent = j.total;
    document.getElementById('cDone').textContent = (j.built||0)+(j.cached||0);
    document.getElementById('cBuilt').textContent = j.built;
    document.getElementById('cFailed').textContent = j.failed;
    document.getElementById('cScan').textContent = j.last_scan ? new Date(j.last_scan*1000).toLocaleString('zh-CN') : '–';
    document.getElementById('cPhase').textContent = j.phase;
    const fails = j.failures||[];
    document.getElementById('fCount').textContent = fails.length;
    const list = document.getElementById('fList');
    list.innerHTML = fails.length ? fails.map(f=>
      '<li><code>'+f.photo_id+'</code>　'+String(f.error||'').split('"').join('')+'</li>').join('')
      : '<li style="border-color:rgba(63,185,80,.3);color:#3fb950">没有失败，全部照片去水印成功 ✓</li>';
  }catch(e){
    document.getElementById('st').className='err';
    document.getElementById('st').textContent='连接失败，服务未运行？';
  }
  setTimeout(load, 3000);
}
load();
</script>
</body>
</html>"""

_NAME_RE = re.compile(r'[^A-Za-z0-9_\-]')
_SESSION_RE = re.compile(r'^[A-Za-z0-9_-]{8,64}$')
_DELIVERY_DAILY_SESSION_LIMIT = 100
_DELIVERY_DAILY_ORDER_LIMIT = 100

# 客户下载时扣摄影师张数额度: 失败原因 -> 面向客户/摄影师的提示
_QUOTA_ERRORS = {
    'no_card':  '摄影师暂无生效卡密，请联系摄影师开通服务',
    'disabled': '摄影师服务已停用，请联系摄影师',
    'expired':  '摄影师卡密已过期，请联系摄影师续费',
    'quota':    '摄影师张数额度已用完，请联系摄影师充值后再下载',
}


def _load_session_secret():
    """摄影师会话签名密钥 (持久化, 重启后已登录会话仍有效)."""
    os.makedirs(DATA_DIR, exist_ok=True)
    if os.path.isfile(SECRET_FILE):
        s = open(SECRET_FILE, 'r').read().strip()
        if s:
            return s.encode('utf-8')
    s = secrets.token_hex(32)
    with open(SECRET_FILE, 'w') as f:
        f.write(s)
    return s.encode('utf-8')


class RequestBodyTooLarge(ValueError):
    pass


def sanitize(name):
    return _NAME_RE.sub('_', str(name))


def project_link(pid, base=''):
    return '%s/#/p/%s' % (base, pid)


# ---------------------------------------------------------------- Excel
def build_xlsx(pid, sel_only=False):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    project = db.get_project(pid)
    photos = db.all_selected(pid) if sel_only else db.list_photos(pid)
    wb = Workbook()
    ws = wb.active
    ws.title = '选片清单'
    headers = ['序号', '照片ID', '导出文件名', '状态', '已选时间']
    ws.append(headers)
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='4472C4')
    sel_fill = PatternFill('solid', fgColor='C6EFCE')
    thin = Border(*[Side(style='thin', color='D9D9D9')] * 4)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal='center')
    for i, ph in enumerate(photos, 1):
        fname = '%03d_%s.jpg' % (i, ph['photo_id'])
        sel_txt = '已选' if ph['selected'] else '未选'
        sel_time = (datetime.fromtimestamp(ph['selected_at']).strftime('%Y-%m-%d %H:%M')
                    if ph.get('selected_at') else '')
        row = [i, ph['photo_id'], fname, sel_txt, sel_time]
        ws.append(row)
        if ph['selected']:
            for c in range(1, len(headers) + 1):
                ws.cell(row=i + 1, column=c).fill = sel_fill
        for c in range(1, len(headers) + 1):
            ws.cell(row=i + 1, column=c).border = thin
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 18
    # 统计页
    ws2 = wb.create_sheet('统计')
    ws2.append(['项目', project['name'] if project else pid])
    ws2.append(['照片总数', len(photos)])
    ws2.append(['已选数量', sum(1 for p in photos if p['selected'])])
    ws2.append(['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws2.column_dimensions['A'].width = 12
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), '%s_%s.xlsx' % (project['name'] if project else pid, datetime.now().strftime('%Y%m%d_%H%M'))


# ---------------------------------------------------------------- 交付系统
def _delivery_code_ok(pid, code):
    """交付码是否对该项目有效 (且交付启用)."""
    if not code or not pid:
        return False
    d = db.get_delivery(pid)
    return bool(d) and d['enabled'] and d['code'] == code


def _order_public(o):
    """订单 -> 客户可见字段."""
    if not o:
        return None
    count = o['count']
    free_count = int(o.get('free_count') or 0)
    return {
        'id': o['id'], 'count': count, 'free_count': free_count,
        'paid_count': max(0, count - free_count),
        'price': o['price'], 'total': o['total'],
        'paid_amount': o['paid_amount'], 'status': o['status'],
        'customer_name': o['customer_name'], 'customer_msg': o['customer_msg'],
        'created_at': o['created_at'],
        'can_download': o['status'] in ('confirmed', 'downloaded'),
    }


def _pricing(d, count):
    """按交付配置计算计价: 返回 (计费单价, 计费张数, 应付金额).
    先扣免费张数, 计费张数>=阶梯门槛时单价减价."""
    price = float(d.get('price') or 0.0)
    free_count = int(d.get('free_count') or 0)
    tier_min = int(d.get('tier_min') or 0)
    tier_discount = float(d.get('tier_discount') or 0.0)
    paid_count = max(0, int(count) - free_count)
    unit = price
    if tier_min > 0 and paid_count >= tier_min and tier_discount > 0:
        unit = max(0.0, price - tier_discount)
    return unit, paid_count, round(unit * paid_count, 2)


def _delivery_job_public(job):
    """Remove internal diagnostics from every customer-visible job response."""
    if not job:
        return job
    public = dict(job)
    public['failures'] = [
        {'photo_id': failure.get('photo_id'), 'code': failure.get('code'),
         'message': '该照片处理失败'}
        for failure in public.get('failures', [])
    ]
    return public


def _delivery_link(d, req_host):
    base = (d['public_base'] or req_host or '').rstrip('/')
    return '%s/d/%s' % (base, d['code'])


# ---------------------------------------------------------------- 工作区扫描
def rescan_workspace(ws_root, owner=None):
    """全工作区扫描: 发现新相册 + 同步每个相册的照片 (仅 listdir/stat, 不解码, 很快).
    owner: 相册归属的摄影师 slug (管理员主工作区 = ADMIN_OWNER)."""
    albums = scanner.find_albums(ws_root)
    for a in albums:
        db.upsert_project(a, owner=owner)
        photos = scanner.scan_project_photos(a['path'])
        disk = []
        for ph in photos:
            ph['project_id'] = a['id']
            ph['key'] = '%s|%s' % (a['id'], ph['photo_id'])
            disk.append(ph)
        db.sync_photos(a['id'], disk)
    return albums


def tenant_ws_root(slug):
    """摄影师租户的工作区根 (与本地工作区同结构: <root>/<user>/<album>/...)."""
    return os.path.join(TENANT_ROOT, slug, 'ws')


def tenant_upload_root(slug):
    """租户工作区下的 project 目录 (上传文件写到这)."""
    return os.path.join(tenant_ws_root(slug), 'project')


def tenant_project_id(slug, album_id):
    """租户相册 ID = "<slug>:<album_id>", 前缀保证不同摄影师永不冲突."""
    return '%s:%s' % (slug, album_id)


def rescan_tenant(slug):
    """扫描某个摄影师租户的上传工作区 (data/tenants/<slug>/ws/project), 归入其名下项目."""
    albums = scanner.find_albums(tenant_upload_root(slug))
    for a in albums:
        pid = tenant_project_id(slug, a['id'])
        db.upsert_project(dict(a, id=pid), owner=slug)
        photos = scanner.scan_project_photos(a['path'])
        disk = []
        for ph in photos:
            ph['project_id'] = pid
            ph['key'] = '%s|%s' % (pid, ph['photo_id'])
            disk.append(ph)
        db.sync_photos(pid, disk)
    return albums


# ---------------------------------------------------------------- HTTP
class Handler(BaseHTTPRequestHandler):
    server_version = 'PixCakeGallery/1.0'

    # ---- helpers
    def setup(self):
        super().setup()
        self.connection.settimeout(20)

    def _query(self, name, default=None):
        qs = parse_qs(urlparse(self.path).query)
        v = qs.get(name)
        return v[0] if v else default

    def _cookie(self, name):
        raw = self.headers.get('Cookie') or ''
        for part in raw.split(';'):
            k, _, v = part.strip().partition('=')
            if k == name:
                return v
        return None

    def _authorized(self):
        tok = self.server.token
        if not tok:
            return True
        return self._cookie('t') == tok or self._query('t') == tok

    # ---- 摄影师租户会话 (cookie g = "<slug>.<hmac>") ----
    def _tenant(self):
        g = self._cookie('g')
        if not g or '.' not in g:
            return None
        slug, _, h = g.rpartition('.')
        expect = self._tenant_hmac(slug)
        if not expect or h != expect:
            return None
        return slug if db.get_photographer(slug) else None

    def _tenant_hmac(self, slug):
        secret = getattr(self.server, 'session_secret', None)
        if not secret:
            return None
        import hashlib
        return hmac.new(secret, ('tenant:' + slug).encode('utf-8'),
                        hashlib.sha256).hexdigest()[:32]

    def _tenant_cookie(self, slug):
        return '%s.%s' % (slug, self._tenant_hmac(slug))

    def _require_tenant(self):
        """返回当前摄影师 slug; 无有效会话则发 401 返回 None."""
        slug = self._tenant()
        if not slug:
            self._json({'error': '摄影师未登录', 'auth': True, 'tenant': True}, 401)
            return None
        return slug

    def _require_tenant_project(self, slug, pid):
        """校验项目归属当前摄影师; 返回项目 dict (含 owner). 非本人 → 403."""
        proj = db.get_project(pid)
        if not proj:
            self._json({'error': '项目不存在'}, 404)
            return None
        if proj.get('owner') and proj['owner'] != slug:
            self._json({'error': '无权访问该项目'}, 403)
            return None
        return proj

    def _json(self, obj, status=200, extra_headers=None):
        body = json.dumps(obj, ensure_ascii=False).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.send_header('Cache-Control', 'no-store')
        if extra_headers:
            for k, v in extra_headers:
                self.send_header(k, v)
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, path, ctype, disposition=None, cache='no-cache'):
        try:
            source = open(path, 'rb') if path else None
        except OSError:
            source = None
        if source is None:
            self.send_response(404)
            self.send_header('Content-Length', '0')
            self.end_headers()
            return False
        try:
            size = os.fstat(source.fileno()).st_size
            self.send_response(200)
            self.send_header('Content-Type', ctype)
            self.send_header('Content-Length', str(size))
            self.send_header('Cache-Control', cache)
            if disposition:
                ascii_name = os.path.basename(path)
                self.send_header('Content-Disposition',
                                 "attachment; filename=\"%s\"; filename*=UTF-8''%s" %
                                 (ascii_name, quote(disposition)))
            self.end_headers()
            while True:
                chunk = source.read(65536)
                if not chunk:
                    break
                self.wfile.write(chunk)
        finally:
            source.close()
        return True

    def _read_json_body(self, max_bytes=1024 * 1024):
        try:
            n = int(self.headers.get('Content-Length') or 0)
        except ValueError:
            n = 0
        if n > max_bytes:
            try:
                self.connection.settimeout(0.25)
                remaining = min(n, max_bytes + 1)
                while remaining:
                    chunk = self.rfile.read(min(65536, remaining))
                    if not chunk:
                        break
                    remaining -= len(chunk)
            except (OSError, TimeoutError):
                pass
            finally:
                self.connection.settimeout(20)
                self.close_connection = True
            raise RequestBodyTooLarge('请求内容过大')
        if n <= 0:
            return {}
        try:
            return json.loads(self.rfile.read(n).decode('utf-8'))
        except (UnicodeDecodeError, ValueError):
            return {}

    # ---- routes
    def do_GET(self):
        try:
            self._handle_get()
        except (BrokenPipeError, ConnectionResetError, ConnectionAbortedError, TimeoutError):
            pass
        except RequestBodyTooLarge as exc:
            self._json({'error': str(exc)}, 413)
        except Exception as e:
            import traceback
            traceback.print_exc()
            try:
                self._json({'error': '服务器内部错误'}, 500)
            except Exception:
                pass

    def do_POST(self):
        try:
            self._handle_post()
        except (BrokenPipeError, ConnectionResetError, ConnectionAbortedError, TimeoutError):
            pass
        except RequestBodyTooLarge as exc:
            self._json({'error': str(exc)}, 413)
        except Exception as e:
            import traceback
            traceback.print_exc()
            try:
                self._json({'error': '服务器内部错误'}, 500)
            except Exception:
                pass

    def _handle_get(self):
        path = urlparse(self.path).path
        if path == '/' or path == '/index.html':
            self._send_file(INDEX_HTML, 'text/html; charset=utf-8')
            return
        if path == '/favicon.ico':
            self._send_file(None, 'image/x-icon')
            return
        if re.match(r'^/d/[^/]+/?$', path) or path == '/delivery.html':
            self._send_file(DELIVERY_HTML, 'text/html; charset=utf-8')
            return
        if path == '/activate.html':
            self._send_file(ACTIVATE_HTML, 'text/html; charset=utf-8')
            return
        m_g = re.match(r'^/g/([A-Za-z0-9_\-]+)/?$', path)
        if m_g:
            # 摄影师站点门户: 访问非本人站点也要展示页面 (页面内判断登录)
            self._send_file(TENANT_HTML, 'text/html; charset=utf-8')
            return
        if path == '/api/tenant/me':
            self._tenant_me_get()
            return
        if path == '/api/tenant/projects':
            self._tenant_projects_get()
            return
        if path == '/api/tenant/project':
            self._tenant_project_get()
            return
        if path == '/api/tenant/delivery':
            self._tenant_delivery_get()
            return
        if path == '/api/tenant/orders':
            self._tenant_orders_get()
            return
        if path == '/api/tenant/stats':
            self._tenant_stats_get()
            return
        if path == '/api/tenant/prewarm':
            self._tenant_prewarm_get()
            return
        if path == '/api/admin/photographers':
            if not self._require_auth():
                return
            self._admin_photographers_get()
            return
        if path == '/api/admin/orders':
            if not self._require_auth():
                return
            self._admin_orders_get()
            return
        if path == '/api/platform/status':
            self._platform_status_get()
            return
        if path == '/api/platform/proxy':
            self._platform_proxy(False)
            return
        if path == '/platform.html' or path == '/platform':
            self._send_file(PLATFORM_HTML, 'text/html; charset=utf-8')
            return
        if path == '/api/license/meta':
            self._license_meta_get()
            return
        if path == '/prewarm':
            if not self._require_auth():
                return
            self._prewarm_page_get()
            return
        if path == '/api/projects':
            if not self._require_auth():
                return
            self.rescan_workspace()
            projs = db.list_projects()
            self._json({'projects': projs})
            return
        if path == '/api/project':
            if not self._require_auth():
                return
            pid = self._query('p') or ''
            self._scan_project_sync(pid)
            proj = db.get_project(pid)
            if not proj:
                self._json({'error': 'project not found'}, 404)
                return
            sel_only = self._query('sel') == '1'
            photos = db.list_photos(pid) if not sel_only else db.all_selected(pid)
            out = []
            for ph in photos:
                thumb = '375' if ph['src_375'] else '3000'
                out.append({
                    'key': ph['key'], 'photo_id': ph['photo_id'],
                    'sort_key': ph['sort_key'], 'selected': ph['selected'],
                    'selected_at': ph['selected_at'],
                    'thumb_url': self._image_url(ph, thumb),
                    'full_url': self._image_url(ph, 3000),
                })
            self._json({'project': proj, 'photos': out})
            return
        if path.startswith('/img/'):
            m = re.match(r'^/img/([^/]+)/', path)
            img_pid = m.group(1) if m else ''
            if not self._authorized():
                if not _delivery_code_ok(img_pid, self._query('k')):
                    self._json({'error': 'auth required', 'auth': True}, 401)
                    return
                if not re.search(r'/375\.jpg$', path):
                    self._json({'error': '客户预览仅提供 375 像素版本'}, 403)
                    return
            self._serve_img(path)
            return
        # ---- 交付端 (客户, 凭交付码, 无管理权限) ----
        if path == '/api/delivery/project':
            self._delivery_project_get()
            return
        if path == '/api/delivery/status':
            self._delivery_status_get()
            return
        if path == '/api/delivery/payqr':
            self._delivery_payqr_get()
            return
        if path == '/api/delivery/download':
            self._delivery_download_get()
            return
        if path == '/api/delivery/export_status':
            self._delivery_export_status_get()
            return
        # ---- 交付端 (管理) ----
        if path == '/api/delivery/setup':
            if not self._require_auth():
                return
            self._delivery_setup_get()
            return
        if path == '/api/delivery/stats':
            if not self._require_auth():
                return
            self._delivery_stats_get()
            return
        if path == '/api/prewarm':
            if not self._require_auth():
                return
            warmer = getattr(self.server, 'warmer', None)
            self._json(warmer.status() if warmer else {'running': False, 'phase': 'idle'})
            return
        if path == '/api/license/admin/list':
            if not self._require_auth():
                return
            self._license_admin_list()
            return
        if path == '/api/delivery/qr':
            if not self._require_auth():
                return
            self._delivery_qr_get()
            return
        if path == '/api/export_xlsx':
            if not self._require_auth():
                return
            pid = self._query('p') or ''
            sel_only = self._query('sel') == '1'
            data, fname = build_xlsx(pid, sel_only)
            self._send_bytes(data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', fname)
            return
        if path == '/api/export_status':
            if not self._require_auth():
                return
            self._export_status_get()
            return
        if path == '/api/export_download':
            if not self._require_auth():
                return
            self._export_download_get()
            return
        self._json({'error': 'not found'}, 404)

    def _handle_post(self):
        path = urlparse(self.path).path
        if path == '/api/select':
            if not self._require_auth():
                return
            body = self._read_json_body()
            keys = body.get('keys') or [body.get('key')]
            sel = 1 if body.get('selected') else 0
            if keys and keys[0]:
                db.set_selected_bulk([k for k in keys if k], sel)
                self._json({'ok': True})
            else:
                self._json({'error': 'no key'}, 400)
            return
        if path == '/api/rename':
            if not self._require_auth():
                return
            body = self._read_json_body()
            pid = body.get('p')
            name = (body.get('name') or '').strip()
            if pid and name:
                db.rename_project(pid, name)
                self._json({'ok': True})
            else:
                self._json({'error': 'p/name required'}, 400)
            return
        if path == '/api/login':
            body = self._read_json_body()
            if self.server.token and body.get('token') == self.server.token:
                self._json({'ok': True}, extra_headers=[('Set-Cookie', 't=%s; Path=/; HttpOnly' % self.server.token)])
            else:
                self._json({'ok': False}, 401)
            return
        # ---- 交付端 (客户) ----
        if path == '/api/delivery/select':
            self._delivery_select_post()
            return
        if path == '/api/delivery/checkout':
            self._delivery_checkout_post()
            return
        if path == '/api/delivery/export':
            self._delivery_export_post()
            return
        if path == '/api/export_zip':
            if not self._require_auth():
                return
            self._export_start_post()
            return
        # ---- 交付端 (管理) ----
        if path == '/api/delivery/setup':
            if not self._require_auth():
                return
            self._delivery_setup_post()
            return
        if path == '/api/delivery/resetcode':
            if not self._require_auth():
                return
            self._delivery_resetcode_post()
            return
        if path == '/api/delivery/confirm':
            if not self._require_auth():
                return
            self._delivery_confirm_post()
            return
        if path == '/api/delivery/uploadqr':
            if not self._require_auth():
                return
            self._delivery_uploadqr_post()
            return
        # ---- 卡密授权 ----
        if path == '/api/license/activate':
            self._license_activate_post()
            return
        if path == '/api/license/verify':
            self._license_verify_handle()
            return
        if path == '/api/license/admin/generate':
            if not self._require_auth():
                return
            self._license_admin_generate()
            return
        if path == '/api/license/admin/toggle':
            if not self._require_auth():
                return
            self._license_admin_toggle()
            return
        if path == '/api/license/admin/extend':
            if not self._require_auth():
                return
            self._license_admin_extend()
            return
        if path == '/api/license/admin/quota':
            if not self._require_auth():
                return
            self._license_admin_quota_post()
            return
        # ---- 管理端: 平台代理 ----
        if path == '/api/platform/config':
            self._platform_config_post()
            return
        if path == '/api/platform/logout':
            self._platform_logout_post()
            return
        if path == '/api/platform/proxy':
            self._platform_proxy(True)
            return
        # ---- 摄影师租户 ----
        if path == '/api/tenant/login':
            self._tenant_login_post()
            return
        if path == '/api/tenant/logout':
            self._tenant_logout_post()
            return
        if path == '/api/tenant/select':
            self._tenant_select_post()
            return
        if path == '/api/tenant/rename':
            self._tenant_rename_post()
            return
        if path == '/api/tenant/delivery':
            self._tenant_delivery_post()
            return
        if path == '/api/tenant/delivery/resetcode':
            self._tenant_resetcode_post()
            return
        if path == '/api/tenant/delivery/confirm':
            self._tenant_confirm_post()
            return
        if path == '/api/tenant/delivery/uploadqr':
            self._tenant_uploadqr_post()
            return
        if path == '/api/tenant/upload':
            self._tenant_upload_post()
            return
        if path == '/api/tenant/scan':
            self._tenant_scan_post()
            return
        # ---- 管理端: 摄影师管理 ----
        if path == '/api/admin/photographers/toggle':
            if not self._require_auth():
                return
            self._admin_photographer_toggle_post()
            return
        if path == '/api/admin/photographers/update':
            if not self._require_auth():
                return
            self._admin_photographer_update_post()
            return
        self._json({'error': 'not found'}, 404)

    def _require_auth(self):
        """返回 True=已授权且(客户端版)已激活; 否则已发送 401/402 并返回 False."""
        if not self._authorized():
            self._json({'error': 'auth required', 'auth': True}, 401)
            return False
        if getattr(self.server, 'license_server', None):
            state = self._license_check()
            if state != 'ok':
                self._json({'error': 'license required', 'license': state}, 402)
                return False
        return True

    # ---- 卡密授权 (license) ----
    def _license_check(self):
        """客户端版: 检查本地授权; 按状态决定是否在线校验. 返回 'ok' 或失败原因.
        blocked/expired/quota 这类服务器可撤销的状态 (解禁用/延期/充值) 定期回连;
        tampered/device 是本地裁定的状态, 保持锁死直到重新激活."""
        st = lic.local_status()
        now = time.time()
        info = st.get('info') or {}
        last = info.get('last_verify') or 0
        if st['state'] == 'ok':
            if now - last > 86400:          # 正常: 每天在线续期一次
                self._license_online_verify()
                st = lic.local_status()
        elif st['state'] in ('blocked', 'expired', 'quota'):
            if now - last > 300:            # 服务器不可达时每 5 分钟才重试一次
                self._license_online_verify()
                st = lic.local_status()
        return st['state']

    def _license_online_verify(self):
        """客户端版: 调授权服务器在线校验, 上报张数卡额度增量, 同步服务器裁决."""
        srv = getattr(self.server, 'license_server', None)
        info = lic.load_local() or {}
        key = (info.get('key') or '').strip()
        if not srv or not key:
            return
        try:
            payload = json.dumps({
                'key': key,
                'machine': lic.machine_fp(),
                'quota_delta': info.get('quota_delta') or 0,
            }).encode('utf-8')
            req = urllib.request.Request(
                srv.rstrip('/') + '/api/license/verify', data=payload,
                headers={'Content-Type': 'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=12) as resp:
                j = json.loads(resp.read().decode('utf-8'))
        except Exception:
            # 授权服务器不可达: 记下时间让冷却生效, 避免每次请求都干等超时
            info = lic.load_local() or {}
            info['last_verify'] = int(time.time())
            lic.save_local(info)
            return
        info = lic.load_local() or {}
        info['last_verify'] = int(time.time())
        if j.get('valid'):
            p = j.get('payload') or {}
            for k in ('key', 'expires_at', 'quota', 'quota_used'):
                if k in p:
                    info[k] = p[k]
            info['quota_delta'] = 0
            info.pop('blocked_at', None)
            info.pop('blocked_reason', None)
        else:
            info['blocked_at'] = int(time.time())
            info['blocked_reason'] = j.get('reason') or 'invalid'
        lic.save_local(info)

    def _license_activate_post(self):
        """POST /api/license/activate. 客户端版: 转发服务器激活+写本地授权;
        授权服务器版: 直接 DB 激活 (供客户端调用). 含同 IP 简单限流."""
        ip = self.client_address[0] if self.client_address else ''
        now = time.time()
        attempts = [t for t in self.server._act_attempts.get(ip, []) if now - t < 30]
        self.server._act_attempts[ip] = attempts
        if len(attempts) >= 5:
            self._json({'error': '尝试过于频繁, 请稍后再试'}, 429)
            return
        attempts.append(now)
        srv = getattr(self.server, 'license_server', None)
        body = self._read_json_body()
        key = (body.get('key') or '').strip().upper()
        if not lic.valid_key(key):
            self._json({'error': '卡密格式不正确'}, 400)
            return
        if srv:
            payload = json.dumps({'key': key, 'machine': lic.machine_fp()}).encode('utf-8')
            try:
                req = urllib.request.Request(
                    srv.rstrip('/') + '/api/license/activate', data=payload,
                    headers={'Content-Type': 'application/json'}, method='POST')
                with urllib.request.urlopen(req, timeout=15) as resp:
                    j = json.loads(resp.read().decode('utf-8'))
            except Exception as exc:
                self._json({'error': '无法连接授权服务器: %s' % exc}, 502)
                return
            if not j.get('ok'):
                self._json({'error': j.get('error') or '激活失败'}, 400)
                return
            info = j.get('payload') or {}
            info['machine'] = lic.machine_fp()
            info['quota_delta'] = 0
            info['last_verify'] = int(time.time())
            info.pop('blocked_at', None)
            info.pop('blocked_reason', None)
            lic.save_local(info)
            self._json({'ok': True, 'payload': info})
            return
        machine = (body.get('machine') or '').strip()
        if not machine:
            self._json({'error': '参数错误'}, 400)
            return
        ok, result = db.activate_license(key, machine)
        if not ok:
            self._json({'error': result}, 400)
            return
        # 创建/绑定摄影师租户 (按机器码唯一)
        pg = db.find_photographer_by_machine(machine)
        if not pg:
            pg = db.create_photographer('摄影师', machine)
        db.bind_card_tenant(key, pg['id'])
        host = self.headers.get('Host', '127.0.0.1')
        result['tenant'] = pg['id']
        result['site_url'] = 'https://%s/g/%s/' % (host, pg['id'])
        self._json({'ok': True, 'payload': result})

    def _license_meta_get(self):
        """GET /api/license/meta (公开). 授权服务器: admin=true;
        客户端: 返回本地授权状态 (含一次到期在线续期尝试), 供前端判断/跳激活页."""
        srv = getattr(self.server, 'license_server', None)
        if not srv:
            self._json({'admin': True, 'state': None, 'info': None})
            return
        state = self._license_check()
        info = lic.load_local() or {}
        self._json({
            'admin': False,
            'state': state,
            'info': {
                'key': info.get('key', ''),
                'plan_name': info.get('plan_name', ''),
                'expires_at': info.get('expires_at') or 0,
                'quota': info.get('quota') or 0,
                'quota_used': info.get('quota_used') or 0,
            } if info else None,
        })

    def _license_verify_handle(self):
        """POST /api/license/verify (授权服务器模式, 供客户端每天续期调用)."""
        body = self._read_json_body()
        key = (body.get('key') or '').strip().upper()
        machine = (body.get('machine') or '').strip()
        try:
            delta = max(0, int(body.get('quota_delta') or 0))
        except (TypeError, ValueError):
            delta = 0
        valid, result = db.verify_license(key, machine, delta)
        if valid:
            self._json({'valid': True, 'payload': result})
        else:
            self._json({'valid': False, 'reason': result})

    def _license_admin_generate(self):
        body = self._read_json_body()
        plan_name = (body.get('plan_name') or '月卡').strip()[:30]
        try:
            duration_days = max(0, int(body.get('duration_days') or 0))
            duration_hours = max(0, int(body.get('duration_hours') or 0))
            quota = max(0, int(body.get('quota') or 0))
            count = max(1, min(100, int(body.get('count') or 1)))
        except (TypeError, ValueError):
            self._json({'error': '数字参数错误'}, 400)
            return
        bind_device = 1 if body.get('bind_device') else 0
        remark = (body.get('remark') or '').strip()[:60]
        # 防御: 套餐名为体验卡但时长全为 0 (参数丢失/编码损坏会落成 0天0小时=永久卡)
        # 强制至少 1 小时, 避免把体验卡错生成永久卡
        if '体验' in plan_name and duration_days == 0 and duration_hours == 0:
            duration_hours = 1
        keys = db.create_license_keys(plan_name, duration_days, quota, bind_device, count,
                                      remark, duration_hours)
        self._json({'ok': True, 'keys': keys, 'count': len(keys)})

    def _license_admin_toggle(self):
        body = self._read_json_body()
        key = (body.get('key') or '').strip().upper()
        status = body.get('status')
        if not key or status not in ('disabled', 'active'):
            self._json({'error': '参数错误'}, 400)
            return
        db.set_license_status(key, status)
        self._json({'ok': True})

    def _license_admin_extend(self):
        body = self._read_json_body()
        key = (body.get('key') or '').strip().upper()
        try:
            days = int(body.get('days') or 0)
        except (TypeError, ValueError):
            days = 0
        if not key or days <= 0:
            self._json({'error': '参数错误'}, 400)
            return
        if not db.extend_license(key, days):
            self._json({'error': '卡密不存在'}, 404)
            return
        self._json({'ok': True})

    def _license_admin_quota_post(self):
        """POST /api/license/admin/quota: 给卡密加/减张数 {key, delta}."""
        body = self._read_json_body()
        key = (body.get('key') or '').strip().upper()
        try:
            delta = int(body.get('delta') or 0)
        except (TypeError, ValueError):
            delta = 0
        if not key or delta == 0:
            self._json({'error': '参数错误'}, 400)
            return
        card = db.add_quota_to_key(key, delta)
        if not card:
            self._json({'error': '卡密不存在'}, 404)
            return
        self._json({'ok': True, 'quota': card['quota'], 'quota_used': card['quota_used']})

    def _license_admin_list(self):
        self._json({'ok': True, 'licenses': db.list_licenses(), 'stats': db.license_stats()})

    # ---------------- 摄影师租户站点 (/g/<slug>/ 与 /api/tenant/*) ----------------
    def _tenant_site_url(self, slug):
        host = self.headers.get('Host', '127.0.0.1')
        return 'https://%s/g/%s/' % (host, slug)

    def _tenant_card_public(self, card):
        if not card:
            return None
        now = int(time.time())
        left_days = None
        if card['expires_at']:
            left_days = max(0, int((card['expires_at'] - now) / 86400))
        quota = card['quota'] or 0
        used = card['quota_used'] or 0
        return {
            'plan_name': card['plan_name'] or '',
            'expires_at': card['expires_at'] or 0,
            'left_days': left_days,
            'unlimited_time': not card['expires_at'],
            'quota': quota,
            'quota_used': used,
            'unlimited_quota': quota == 0,
            'remaining': (quota - used) if quota else -1,
            'status': card['status'],
        }

    def _tenant_login_post(self):
        """POST /api/tenant/login: 摄影师用卡密登录自己的站点.
        body: {key, machine?}. machine 客户端上传时携带以做绑机校验."""
        body = self._read_json_body()
        key = (body.get('key') or '').strip().upper()
        machine = (body.get('machine') or '').strip()
        if not lic.valid_key(key):
            self._json({'error': '卡密格式不正确'}, 400)
            return
        card = db.get_license(key)
        if not card:
            self._json({'error': '卡密不存在'}, 404)
            return
        if card['status'] == 'disabled':
            self._json({'error': '卡密已被禁用，请联系管理员'}, 403)
            return
        if card['status'] == 'unused':
            self._json({'error': '卡密尚未激活，请先在客户端软件中激活'}, 403)
            return
        if card['status'] == 'expired' or (card['expires_at'] and time.time() > card['expires_at']):
            if card['status'] != 'expired':
                db.set_license_status(key, 'expired')
            self._json({'error': '卡密已过期，请联系管理员购买新卡密'}, 403)
            return
        if card['bind_device'] and card['bound_fp'] and machine and machine != card['bound_fp']:
            self._json({'error': '卡密已绑定其他设备'}, 403)
            return
        tenant = card['tenant']
        if not tenant:
            pg = db.find_photographer_by_machine(machine) if machine else None
            pg = pg or db.create_photographer('摄影师', machine)
            tenant = pg['id']
            db.bind_card_tenant(key, tenant)
        pg = db.get_photographer(tenant)
        if not pg:
            pg = db.create_photographer('摄影师', machine)
            tenant = pg['id']
            db.bind_card_tenant(key, tenant)
        if pg['status'] == 'disabled':
            self._json({'error': '站点已被管理员停用'}, 403)
            return
        db.update_photographer(tenant, updated_at=int(time.time()))
        active = db.tenant_active_card(tenant)
        self._json({'ok': True, 'tenant': {
            'id': pg['id'], 'name': pg['name'], 'status': pg['status'],
            'admin_contact': pg['admin_contact'] or '',
            'site_url': self._tenant_site_url(pg['id']),
        }, 'card': self._tenant_card_public(active or card)},
            extra_headers=[('Set-Cookie', 'g=%s; Path=/; HttpOnly' % self._tenant_cookie(tenant))])

    def _tenant_logout_post(self):
        self._json({'ok': True}, extra_headers=[
            ('Set-Cookie', 'g=; Path=/; Max-Age=0; HttpOnly')])

    def _tenant_me_get(self):
        """GET /api/tenant/me: 摄影师本人信息 + 卡密状态 + 站点概况."""
        slug = self._require_tenant()
        if not slug:
            return
        pg = db.get_photographer(slug)
        if not pg:
            self._json({'error': '摄影师不存在'}, 404)
            return
        card = db.tenant_active_card(slug)
        projs = db.list_projects_for(slug)
        order_count = db.count_orders_for_owner(slug)
        photo_count = db.count_photos_for_owner(slug)
        self._json({
            'tenant': {
                'id': pg['id'], 'name': pg['name'], 'contact': pg['contact'],
                'status': pg['status'], 'admin_contact': pg['admin_contact'] or '',
                'site_url': self._tenant_site_url(pg['id']),
            },
            'card': self._tenant_card_public(card),
            'stats': {'projects': len(projs), 'photos': photo_count, 'orders': order_count},
        })

    def _tenant_projects_get(self):
        slug = self._require_tenant()
        if not slug:
            return
        rescan_tenant(slug)          # 扫描租户工作区 (上传的新相册实时出现)
        projs = []
        for p in db.list_projects_for(slug):
            d = db.get_delivery(p['id'])
            projs.append({**p, 'code': d['code'] if d else '',
                          'delivery_enabled': bool(d and d['enabled'])})
        self._json({'projects': projs})

    def _tenant_project_get(self):
        slug = self._require_tenant()
        if not slug:
            return
        pid = self._query('p') or ''
        proj = self._require_tenant_project(slug, pid)
        if not proj:
            return
        self._scan_project_sync(pid)
        photos = db.list_photos(pid)
        out = []
        for ph in photos:
            thumb = '375' if ph['src_375'] else '3000'
            out.append({
                'key': ph['key'], 'photo_id': ph['photo_id'],
                'sort_key': ph['sort_key'], 'selected': ph['selected'],
                'selected_at': ph['selected_at'],
                'thumb_url': self._image_url(ph, thumb),
                'full_url': self._image_url(ph, 3000),
            })
        self._json({'project': proj, 'photos': out})

    def _tenant_select_post(self):
        slug = self._require_tenant()
        if not slug:
            return
        body = self._read_json_body()
        keys = body.get('keys') or [body.get('key')]
        sel = 1 if body.get('selected') else 0
        if keys and keys[0]:
            pid = keys[0].split('|')[0]
            if not self._require_tenant_project(slug, pid):
                return
            db.set_selected_bulk([k for k in keys if k], sel)
            self._json({'ok': True})
        else:
            self._json({'error': 'no key'}, 400)

    def _tenant_rename_post(self):
        slug = self._require_tenant()
        if not slug:
            return
        body = self._read_json_body()
        pid = body.get('p')
        name = (body.get('name') or '').strip()
        if pid and name and self._require_tenant_project(slug, pid):
            db.rename_project(pid, name)
            self._json({'ok': True})
        else:
            self._json({'error': 'p/name required'}, 400)

    def _tenant_delivery_get(self):
        slug = self._require_tenant()
        if not slug:
            return
        pid = self._query('p') or ''
        if not self._require_tenant_project(slug, pid):
            return
        d = db.ensure_delivery(pid)
        host = self.headers.get('Host', '127.0.0.1:%d' % self.server.server_port)
        self._json({'config': {
            'project_id': d['project_id'], 'code': d['code'], 'title': d['title'],
            'price': d['price'] or 0.0,
            'free_count': int(d.get('free_count') or 0),
            'tier_min': int(d.get('tier_min') or 0),
            'tier_discount': float(d.get('tier_discount') or 0.0),
            'enabled': d['enabled'],
            'public_base': d['public_base'] or '',
            'has_pay_qr': bool(d['pay_qr_path']),
        }, 'link': _delivery_link(d, host)})

    def _tenant_delivery_post(self):
        slug = self._require_tenant()
        if not slug:
            return
        body = self._read_json_body()
        pid = body.get('p') or self._query('p') or ''
        if not self._require_tenant_project(slug, pid):
            return
        db.ensure_delivery(pid)
        kw = {}
        if 'title' in body:
            kw['title'] = str(body.get('title') or '').strip()[:100]
        if 'price' in body:
            try:
                price = float(body.get('price') or 0)
                kw['price'] = price if math.isfinite(price) and 0 <= price <= 1000000 else 0.0
            except (TypeError, ValueError):
                kw['price'] = 0.0
        if 'enabled' in body:
            kw['enabled'] = 1 if body.get('enabled') else 0
        if 'public_base' in body:
            kw['public_base'] = str(body.get('public_base') or '').strip().rstrip('/')[:500]
        if 'free_count' in body:
            try:
                fc = int(body.get('free_count') or 0)
                kw['free_count'] = fc if 0 <= fc <= 10000 else 0
            except (TypeError, ValueError):
                kw['free_count'] = 0
        if 'tier_min' in body:
            try:
                tm = int(body.get('tier_min') or 0)
                kw['tier_min'] = tm if 0 <= tm <= 10000 else 0
            except (TypeError, ValueError):
                kw['tier_min'] = 0
        if 'tier_discount' in body:
            try:
                td = float(body.get('tier_discount') or 0)
                kw['tier_discount'] = td if math.isfinite(td) and 0 <= td <= 1000000 else 0.0
            except (TypeError, ValueError):
                kw['tier_discount'] = 0.0
        db.update_delivery(pid, **kw)
        d = db.get_delivery(pid)
        host = self.headers.get('Host', '127.0.0.1:%d' % self.server.server_port)
        self._json({'ok': True, 'link': _delivery_link(d, host)})

    def _tenant_resetcode_post(self):
        slug = self._require_tenant()
        if not slug:
            return
        body = self._read_json_body()
        pid = body.get('p') or self._query('p') or ''
        if not self._require_tenant_project(slug, pid):
            return
        d = db.reset_delivery_code(pid)
        host = self.headers.get('Host', '127.0.0.1:%d' % self.server.server_port)
        self._json({'ok': True, 'code': d['code'], 'link': _delivery_link(d, host)})

    def _tenant_confirm_post(self):
        slug = self._require_tenant()
        if not slug:
            return
        body = self._read_json_body()
        o = db.get_order(body.get('order_id')) if body.get('order_id') else None
        if not o or o['project_id'] not in {p['id'] for p in db.list_projects_for(slug)}:
            self._json({'error': '订单不存在'}, 404)
            return
        if body.get('agree'):
            db.confirm_order(o['id'])
        elif body.get('reject'):
            db.delete_order(o['id'])
        self._json({'ok': True, 'order': _order_public(db.get_order(o['id']))})

    def _tenant_uploadqr_post(self):
        slug = self._require_tenant()
        if not slug:
            return
        body = self._read_json_body(max_bytes=8 * 1024 * 1024)
        pid = body.get('p') or ''
        if not self._require_tenant_project(slug, pid):
            return
        data_url = body.get('data_url') or ''
        if not data_url.startswith('data:image/'):
            self._json({'error': '需要 base64 图片'}, 400)
            return
        m = re.match(r'^data:image/(\w+);base64,(.*)$', data_url, re.S)
        if not m:
            self._json({'error': '图片格式不支持'}, 400)
            return
        import base64
        ext = 'png' if m.group(1).lower() == 'png' else 'jpg'
        try:
            raw = base64.b64decode(m.group(2))
        except Exception:
            self._json({'error': '图片解码失败'}, 400)
            return
        if len(raw) > 5 * 1024 * 1024:
            self._json({'error': '图片不能超过 5 MB'}, 400)
            return
        os.makedirs(QR_DIR, exist_ok=True)
        path = os.path.join(QR_DIR, sanitize(pid) + '.' + ext)
        with open(path, 'wb') as f:
            f.write(raw)
        db.ensure_delivery(pid)
        db.update_delivery(pid, pay_qr_path=path)
        self._json({'ok': True, 'has_pay_qr': True})

    def _tenant_orders_get(self):
        slug = self._require_tenant()
        if not slug:
            return
        pids = [p['id'] for p in db.list_projects_for(slug)]
        orders = []
        for pid in pids:
            for o in db.list_orders(pid):
                orders.append({**dict(o), 'project_id': pid})
        orders.sort(key=lambda x: x.get('created_at') or 0, reverse=True)
        self._json({'orders': orders})

    def _tenant_stats_get(self):
        slug = self._require_tenant()
        if not slug:
            return
        card = db.tenant_active_card(slug)
        self._json({'card': self._tenant_card_public(card),
                    'stats': db.tenant_stats(slug)})

    def _tenant_upload_post(self):
        """POST /api/tenant/upload: 客户端分批上传像素蛋糕工作区文件到自己的租户.
        body: {rel_path, data_b64, offset?, machine?}. offset>0 续传."""
        slug = self._require_tenant()
        if not slug:
            return
        body = self._read_json_body(max_bytes=24 * 1024 * 1024)
        card = db.tenant_active_card(slug)
        if not card:
            self._json({'error': '无生效卡密，请联系管理员购买'}, 402)
            return
        if card['status'] != 'active':
            self._json({'error': '卡密不可用'}, 402)
            return
        if card['expires_at'] and time.time() > card['expires_at']:
            db.set_license_status(card['key'], 'expired')
            self._json({'error': '卡密已过期，请联系管理员续费'}, 402)
            return
        machine = (body.get('machine') or '').strip()
        if card['bind_device'] and card['bound_fp'] and machine and machine != card['bound_fp']:
            self._json({'error': '卡密已绑定其他设备'}, 403)
            return
        rel = (body.get('rel_path') or '').strip().replace('\\', '/')
        if (not rel or rel.startswith('/') or '..' in rel or ':' in rel
                or not re.match(r'^[\w\-一-鿿\./\s]+$', rel)):
            self._json({'error': '非法路径'}, 400)
            return
        import base64
        try:
            raw = base64.b64decode(body.get('data_b64') or '')
        except Exception:
            self._json({'error': '数据解码失败'}, 400)
            return
        if not raw:
            self._json({'error': '空数据'}, 400)
            return
        if len(raw) > 16 * 1024 * 1024:
            self._json({'error': '单块数据过大'}, 400)
            return
        root = os.path.realpath(tenant_upload_root(slug))
        os.makedirs(root, exist_ok=True)
        dest = os.path.realpath(os.path.join(root, rel))
        if not dest.startswith(root + os.sep):
            self._json({'error': '非法路径'}, 400)
            return
        try:
            offset = int(body.get('offset') or 0)
        except (TypeError, ValueError):
            offset = 0
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        with open(dest, 'wb' if offset == 0 else 'ab') as f:
            if offset > 0:
                f.seek(offset)
            f.write(raw)
        self._json({'ok': True, 'size': os.path.getsize(dest)})

    def _tenant_scan_post(self):
        """POST /api/tenant/scan: 上传完后让服务器扫描租户工作区 + 触发去水印."""
        slug = self._require_tenant()
        if not slug:
            return
        rescan_tenant(slug)
        tw = getattr(self.server, 'tenant_warmer', None)
        if tw:
            tw.trigger(slug)
        projs = []
        for p in db.list_projects_for(slug):
            d = db.get_delivery(p['id'])
            projs.append({**p, 'code': d['code'] if d else ''})
        self._json({'ok': True, 'projects': projs})

    def _tenant_prewarm_get(self):
        slug = self._require_tenant()
        if not slug:
            return
        tw = getattr(self.server, 'tenant_warmer', None)
        self._json({'ok': True, 'prewarm': tw.status(slug) if tw else {}})

    # ---------------- 管理端: 摄影师管理 ----------------
    def _admin_photographers_get(self):
        rows = db.list_photographers()
        out = []
        for pg in rows:
            card = db.tenant_active_card(pg['id'])
            out.append({
                'id': pg['id'], 'name': pg['name'], 'contact': pg['contact'],
                'machine_fp': pg['machine_fp'], 'status': pg['status'],
                'admin_contact': pg['admin_contact'] or '',
                'created_at': pg['created_at'],
                'card': self._tenant_card_public(card),
                'projects': len(db.list_projects_for(pg['id'])),
                'orders': db.count_orders_for_owner(pg['id']),
            })
        self._json({'ok': True, 'photographers': out})

    def _admin_photographer_update_post(self):
        body = self._read_json_body()
        slug = (body.get('id') or '').strip()
        kw = {}
        if 'name' in body:
            kw['name'] = str(body['name'] or '')[:60]
        if 'contact' in body:
            kw['contact'] = str(body['contact'] or '')[:200]
        if 'admin_contact' in body:
            kw['admin_contact'] = str(body['admin_contact'] or '')[:200]
        if not db.update_photographer(slug, **kw):
            self._json({'error': '摄影师不存在'}, 404)
            return
        self._json({'ok': True})

    def _admin_photographer_toggle_post(self):
        body = self._read_json_body()
        slug = (body.get('id') or '').strip()
        status = body.get('status')
        if status not in ('active', 'disabled'):
            self._json({'error': '参数错误'}, 400)
            return
        db.set_photographer_status(slug, status)
        self._json({'ok': True})

    def _admin_orders_get(self):
        """GET /api/admin/orders: 全平台订单总览."""
        rows = db.list_all_orders(limit=300)
        out = []
        for o in rows:
            try:
                photo_ids = json.loads(o['photo_ids'] or '[]')
            except ValueError:
                photo_ids = []
            out.append({
                'id': o['id'], 'owner': o.get('owner') or '',
                'project_id': o['project_id'], 'project_name': o.get('project_name') or '',
                'code': o.get('code') or '', 'session': o['session'],
                'count': o['count'] or len(photo_ids), 'total': o['total'] or 0,
                'paid_amount': o['paid_amount'] or 0,
                'customer_name': o['customer_name'] or '',
                'status': o['status'],
                'created_at': o['created_at'], 'confirmed_at': o['confirmed_at'],
            })
        self._json({'ok': True, 'orders': out})

    # ---------------- 管理端平台代理 (/api/platform/*) ----------------
    # 管理端本地跑 gallery 时, 用这个把平台管理请求转发到中央服务器 (管理 token 存在本地).
    def _platform_cfg(self):
        try:
            with open(PLATFORM_FILE, encoding='utf-8') as f:
                cfg = json.load(f)
            if isinstance(cfg, dict):
                return cfg
        except Exception:
            pass
        return {}

    def _platform_cfg_save(self, cfg):
        os.makedirs(DATA_DIR, exist_ok=True)
        tmp = PLATFORM_FILE + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, ensure_ascii=False)
        os.replace(tmp, PLATFORM_FILE)

    def _platform_status_get(self):
        cfg = self._platform_cfg()
        self._json({'configured': bool(cfg.get('server') and cfg.get('token')),
                    'server': cfg.get('server') or '', 'token': cfg.get('token') or ''})

    def _platform_config_post(self):
        body = self._read_json_body()
        server = (body.get('server') or '').strip().rstrip('/')
        token = (body.get('token') or '').strip()
        if not re.match(r'^https?://', server) or not token:
            self._json({'error': '服务器地址或令牌无效'}, 400)
            return
        # 先验证连接
        import urllib.error
        req = urllib.request.Request(server + '/api/admin/photographers',
                                     headers={'Cookie': 't=%s' % token})
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                resp.read()
        except urllib.error.HTTPError as exc:
            self._json({'error': '令牌验证失败 (HTTP %d)' % exc.code}, 400)
            return
        except Exception as exc:
            self._json({'error': '无法连接平台服务器: %s' % exc}, 502)
            return
        self._platform_cfg_save({'server': server, 'token': token})
        self._json({'ok': True, 'server': server})

    def _platform_logout_post(self):
        try:
            os.remove(PLATFORM_FILE)
        except OSError:
            pass
        self._json({'ok': True})

    def _platform_proxy(self, is_post):
        """把 /api/platform/proxy 转发到中央服务器对应的 /api/* (携带管理 token)."""
        cfg = self._platform_cfg()
        if not cfg.get('server') or not cfg.get('token'):
            self._json({'error': '未配置平台连接'}, 400)
            return
        if is_post:
            body = self._read_json_body()
            path = (body.get('path') or '').strip()
            payload = json.dumps(body.get('body') or {}).encode('utf-8')
        else:
            body = {}
            path = (self._query('path') or '').strip()
            payload = None
        if not path.startswith('/api/'):
            self._json({'error': '非法路径'}, 400)
            return
        import urllib.error
        req = urllib.request.Request(
            cfg['server'].rstrip('/') + path, data=payload, method='POST' if is_post else 'GET',
            headers={'Content-Type': 'application/json', 'Cookie': 't=%s' % cfg['token']})
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                raw = resp.read().decode('utf-8')
            self._json(json.loads(raw))
        except urllib.error.HTTPError as exc:
            raw = exc.read().decode('utf-8', 'replace')
            try:
                self._json(json.loads(raw), exc.code)
            except Exception:
                self._json({'error': '平台服务器错误 (HTTP %d)' % exc.code}, exc.code)
        except Exception as exc:
            self._json({'error': '无法连接平台服务器: %s' % exc}, 502)

    def _scan_project_sync(self, pid):
        proj = db.get_project(pid)
        if not proj:
            return
        photos = scanner.scan_project_photos(proj['path'])
        disk = []
        for ph in photos:
            ph['project_id'] = pid
            ph['key'] = '%s|%s' % (pid, ph['photo_id'])
            disk.append(ph)
        db.sync_photos(pid, disk)

    def rescan_workspace(self):
        rescan_workspace(self.server.ws_root)

    def _serve_img(self, path):
        # /img/<pid>/<photo_id>/<size>.jpg
        m = re.match(r'^/img/([^/]+)/([^/]+)/(\d+)\.jpg$', path)
        if not m:
            self._json({'error': 'bad img path'}, 400)
            return
        pid, photo_id, size = m.group(1), m.group(2), int(m.group(3))
        if size not in (375, 3000):
            self._json({'error': 'bad size'}, 400)
            return
        ph = db.get_photo('%s|%s' % (pid, photo_id))
        if not ph or not ph['on_disk']:
            self.send_response(404)
            self.send_header('Content-Length', '0')
            self.end_headers()
            return
        try:
            cached = self.server.image_service.ensure(ph, size)
        except ImageServiceError as exc:
            failure = exc.to_dict()
            if not self._authorized():
                failure.pop('details', None)
            self._json({'error': str(exc), 'failure': failure}, 422)
            return
        self._send_file(cached, 'image/jpeg', cache='public, max-age=31536000, immutable')

    def _image_url(self, photo, size, delivery_code=None):
        signature = self.server.image_service.source_signature(photo)
        params = []
        if delivery_code:
            params.append('k=%s' % quote(delivery_code))
        params.extend([
            'e=%s' % signature['refined'].get('mtime_ns', 0),
            'o=%s' % signature['original'].get('mtime_ns', 0),
            'a=%s' % self.server.image_service.algorithm_hash[:16],
        ])
        return '/img/%s/%s/%s.jpg?%s' % (
            photo['project_id'], photo['photo_id'], size, '&'.join(params)
        )

    def _export_key(self, prefix, pid, photos):
        revisions = [
            '%s:%s' % (photo['photo_id'], self.server.image_service.revision(photo))
            for photo in photos
        ]
        return '%s|%s|%s' % (prefix, pid, '|'.join(revisions))

    def _export_start_post(self):
        body = self._read_json_body()
        pid = body.get('p') or ''
        selected_only = body.get('selected_only', True) is not False
        self._scan_project_sync(pid)
        if not db.get_project(pid):
            self._json({'error': '项目不存在'}, 404)
            return
        photos = db.all_selected(pid) if selected_only else db.list_photos(pid)
        job = self.server.export_manager.start(
            pid,
            photo_ids=[photo['photo_id'] for photo in photos],
            filename_prefix='选中照片' if selected_only else '全部照片',
            context={'scope': 'admin'},
            dedupe_key=self._export_key('admin-selected' if selected_only else 'admin-all', pid, photos),
        )
        self._json(job, 202)

    def _export_status_get(self):
        job_id = self._query('id') or ''
        if not self.server.export_manager.matches(job_id, scope='admin'):
            self._json({'error': '导出任务不存在'}, 404)
            return
        self._json(self.server.export_manager.status(job_id))

    def _export_download_get(self):
        job_id = self._query('id') or ''
        if not self.server.export_manager.matches(job_id, scope='admin'):
            self._json({'error': '导出任务不存在'}, 404)
            return
        result = self.server.export_manager.download(job_id)
        if not result:
            self._json({'error': '导出文件尚未生成'}, 409)
            return
        path, filename = result
        self._send_file(path, 'application/zip', filename, cache='no-store')

    # ---------------- 交付端 (客户, 凭交付码) ----------------
    def _delivery_project_get(self):
        d = db.get_delivery_by_code(self._query('k') or '')
        if not d:
            self._json({'error': '无效的交付码'}, 403)
            return
        pid = d['project_id']
        self._scan_project_sync(pid)                # 实时发现新精修
        session = self._query('s') or ''
        photos = []
        for ph in db.list_photos(pid):
            photos.append({
                'photo_id': ph['photo_id'],
                'thumb_url': self._image_url(ph, 375, d['code']),
                'full_url': self._image_url(ph, 375, d['code']),
            })
        proj = db.get_project(pid)
        sel = db.get_selections(pid, session) if session else {}
        order = _order_public(db.get_active_order(pid, session)) if session else None
        self._json({
            'project': {
                'id': pid,
                'title': d['title'] or (proj['name'] if proj else pid),
                'album': proj['album_id'] if proj else '',
                'price': d['price'] or 0.0,
                'free_count': int(d.get('free_count') or 0),
                'tier_min': int(d.get('tier_min') or 0),
                'tier_discount': float(d.get('tier_discount') or 0.0),
                'enabled': d['enabled'],
                'photo_count': len(photos),
            },
            'photos': photos,
            'selections': sel,
            'order': order,
        })

    def _delivery_status_get(self):
        d = db.get_delivery_by_code(self._query('k') or '')
        if not d:
            self._json({'error': '无效的交付码'}, 403)
            return
        session = self._query('s') or ''
        o = db.get_order(self._query('order_id')) if self._query('order_id') else None
        if not o or o['project_id'] != d['project_id'] or o['session'] != session:
            self._json({'error': '订单不存在'}, 404)
            return
        out = _order_public(o)
        cfg = db.get_delivery(d['project_id'])
        if cfg and cfg.get('pay_qr_path') and (o['price'] or 0) > 0:
            out['pay_qr_url'] = '/api/delivery/payqr?k=%s' % d['code']
        self._json(out)

    def _delivery_payqr_get(self):
        d = db.get_delivery_by_code(self._query('k') or '')
        if not d or not d.get('pay_qr_path') or not os.path.isfile(d['pay_qr_path']):
            self.send_response(404)
            self.send_header('Content-Length', '0')
            self.end_headers()
            return
        ext = os.path.splitext(d['pay_qr_path'])[1].lower().lstrip('.')
        self._send_file(d['pay_qr_path'], 'image/png' if ext == 'png' else 'image/jpeg',
                        cache='public, max-age=604800')

    def _delivery_download_get(self):
        d = db.get_delivery_by_code(self._query('k') or '')
        if not d:
            self._json({'error': '无效的交付码'}, 403)
            return
        session = self._query('s') or ''
        o = db.get_order(self._query('order_id')) if self._query('order_id') else None
        if not o or o['project_id'] != d['project_id'] or o['session'] != session:
            self._json({'error': '订单不存在'}, 404)
            return
        if o['status'] not in ('confirmed', 'downloaded'):
            self._json({'error': '摄影师尚未确认，暂不能下载'}, 403)
            return
        job_id = self._query('job_id') or ''
        if not self.server.export_manager.matches(
                job_id, scope='delivery', order_id=o['id'], session=session):
            self._json({'error': '下载任务不存在'}, 404)
            return
        # 首次下载: 服务器强校验并扣摄影师张数额度 (重复下载不重复扣)
        if o['status'] == 'confirmed':
            owner = db.project_owner(d['project_id'])
            if owner and owner != ADMIN_OWNER:
                try:
                    count = max(1, int(o['count'] or 0))
                except (TypeError, ValueError):
                    count = len(json.loads(o['photo_ids'] or '[]')) or 1
                ok, reason = db.consume_quota(owner, count)
                if not ok:
                    self._json({'error': _QUOTA_ERRORS.get(reason, '张数额度不足')}, 402)
                    return
        result = self.server.export_manager.download(job_id)
        if not result:
            self._json({'error': '照片仍在处理中'}, 409)
            return
        path, filename = result
        if self._send_file(path, 'application/zip', filename, cache='no-store'):
            db.mark_order_downloaded(o['id'])

    def _delivery_export_status_get(self):
        d = db.get_delivery_by_code(self._query('k') or '')
        session = self._query('s') or ''
        try:
            order_id = int(self._query('order_id') or 0)
        except ValueError:
            order_id = 0
        o = db.get_order(order_id) if order_id else None
        if not d or not o or o['project_id'] != d['project_id'] or o['session'] != session:
            self._json({'error': '订单不存在'}, 404)
            return
        job_id = self._query('id') or ''
        if not self.server.export_manager.matches(
                job_id, scope='delivery', order_id=o['id'], session=session):
            self._json({'error': '导出任务不存在'}, 404)
            return
        self._json(_delivery_job_public(self.server.export_manager.status(job_id)))

    def _delivery_export_post(self):
        body = self._read_json_body()
        d = db.get_delivery_by_code(body.get('k') or '')
        session = body.get('session') or ''
        try:
            order_id = int(body.get('order_id') or 0)
        except (TypeError, ValueError):
            order_id = 0
        o = db.get_order(order_id) if order_id else None
        if not d or not o or o['project_id'] != d['project_id'] or o['session'] != session:
            self._json({'error': '订单不存在'}, 404)
            return
        if o['status'] not in ('confirmed', 'downloaded'):
            self._json({'error': '摄影师尚未确认，暂不能下载'}, 403)
            return
        try:
            photo_ids = json.loads(o['photo_ids'] or '[]')
        except ValueError:
            photo_ids = []
        if not photo_ids:
            self._json({'error': '订单无照片'}, 400)
            return
        # 张数卡: 导出前先预检摄影师额度 (不够则拒绝, 不让客户白等)
        owner = db.project_owner(d['project_id'])
        if owner and owner != ADMIN_OWNER:
            try:
                count = max(1, int(o['count'] or 0))
            except (TypeError, ValueError):
                count = len(photo_ids)
            ok, reason = db.check_quota(owner, count)
            if not ok:
                self._json({'error': _QUOTA_ERRORS.get(reason, '张数额度不足')}, 402)
                return
        photos = []
        for photo_id in photo_ids:
            photo = db.get_photo('%s|%s' % (d['project_id'], photo_id))
            if photo and photo.get('on_disk'):
                photos.append(photo)
        job = self.server.export_manager.start(
            d['project_id'],
            photo_ids=photo_ids,
            filename_prefix='选片下载',
            context={'scope': 'delivery', 'order_id': o['id'], 'session': session},
            dedupe_key=self._export_key('delivery-%s' % o['id'], d['project_id'], photos),
        )
        self._json(_delivery_job_public(job), 202)

    def _delivery_select_post(self):
        body = self._read_json_body()
        d = db.get_delivery_by_code(body.get('k') or self._query('k') or '')
        if not d:
            self._json({'error': '无效的交付码'}, 403)
            return
        session = body.get('session') or ''
        if not _SESSION_RE.fullmatch(session):
            self._json({'error': '缺少会话'}, 400)
            return
        # 已有待确认订单时禁止改选, 防止金额错乱
        if db.get_active_order(d['project_id'], session):
            self._json({'error': '已有提交的订单，无法再修改选片'}, 400)
            return
        clean = []
        items = body.get('items') or []
        if not isinstance(items, list) or len(items) > 500:
            self._json({'error': '选片数据过多'}, 400)
            return
        valid_ids = {photo['photo_id'] for photo in db.list_photos(d['project_id'])}
        for it in items:
            if not isinstance(it, dict):
                continue
            phid = str(it.get('photo_id') or '')
            if phid in valid_ids:
                clean.append({'photo_id': phid, 'selected': it.get('selected'),
                              'note': str(it.get('note') or '')[:500]})
        since = int(time.time()) - 24 * 3600
        if (not db.delivery_session_exists(d['project_id'], session) and
                db.count_delivery_sessions_since(d['project_id'], since) >=
                _DELIVERY_DAILY_SESSION_LIMIT):
            self._json({'error': '今日新会话数量已达上限，请稍后重试'}, 429)
            return
        db.save_selections_bulk(d['project_id'], session, clean)
        self._json({'ok': True, 'selections': db.get_selections(d['project_id'], session)})

    def _delivery_checkout_post(self):
        body = self._read_json_body()
        d = db.get_delivery_by_code(body.get('k') or self._query('k') or '')
        if not d:
            self._json({'error': '无效的交付码'}, 403)
            return
        session = body.get('session') or ''
        if not _SESSION_RE.fullmatch(session):
            self._json({'error': '缺少会话'}, 400)
            return
        pid = d['project_id']
        order = db.get_active_order(pid, session)
        if order:
            self._json({'ok': True, 'order': _order_public(order)})
            return
        sel = db.get_selections(pid, session)
        valid_ids = {photo['photo_id'] for photo in db.list_photos(pid)}
        photo_ids = [phid for phid, st in sel.items()
                     if st.get('selected') and phid in valid_ids]
        if not photo_ids:
            self._json({'error': '请先选择要下载的照片'}, 400)
            return
        if db.count_orders_since(pid, int(time.time()) - 24 * 3600) >= _DELIVERY_DAILY_ORDER_LIMIT:
            self._json({'error': '今日订单数量已达上限，请稍后重试'}, 429)
            return
        price, _paid_count, total = _pricing(d, len(photo_ids))
        free_count = int(d.get('free_count') or 0)
        paid = body.get('paid_amount')
        try:
            paid = float(paid) if paid not in (None, '') else total
        except (TypeError, ValueError):
            paid = total
        if not math.isfinite(paid) or paid < 0 or paid > 100000000:
            paid = total
        oid = db.create_order(pid, session, photo_ids, price, free_count, total, paid,
                              str(body.get('customer_name') or '')[:50],
                              str(body.get('customer_msg') or '')[:500])
        self._json({'ok': True, 'order': _order_public(db.get_order(oid))})

    # ---------------- 交付端 (管理) ----------------
    def _delivery_setup_get(self):
        pid = self._query('p') or ''
        if not pid:
            self._json({'error': '缺少 p'}, 400)
            return
        d = db.ensure_delivery(pid)
        host = self.headers.get('Host', '127.0.0.1:%d' % self.server.server_port)
        self._json({
            'config': {
                'project_id': d['project_id'], 'code': d['code'], 'title': d['title'],
                'price': d['price'] or 0.0,
                'free_count': int(d.get('free_count') or 0),
                'tier_min': int(d.get('tier_min') or 0),
                'tier_discount': float(d.get('tier_discount') or 0.0),
                'enabled': d['enabled'],
                'public_base': d['public_base'] or '',
                'has_pay_qr': bool(d['pay_qr_path']),
            },
            'link': _delivery_link(d, host),
        })

    def _delivery_setup_post(self):
        body = self._read_json_body()
        pid = body.get('p') or self._query('p') or ''
        if not pid:
            self._json({'error': '缺少 p'}, 400)
            return
        db.ensure_delivery(pid)
        kw = {}
        if 'title' in body:
            kw['title'] = str(body.get('title') or '').strip()[:100]
        if 'price' in body:
            try:
                price = float(body.get('price') or 0)
                kw['price'] = price if math.isfinite(price) and 0 <= price <= 1000000 else 0.0
            except (TypeError, ValueError):
                kw['price'] = 0.0
        if 'enabled' in body:
            kw['enabled'] = 1 if body.get('enabled') else 0
        if 'public_base' in body:
            kw['public_base'] = str(body.get('public_base') or '').strip().rstrip('/')[:500]
        if 'free_count' in body:
            try:
                fc = int(body.get('free_count') or 0)
                kw['free_count'] = fc if 0 <= fc <= 10000 else 0
            except (TypeError, ValueError):
                kw['free_count'] = 0
        if 'tier_min' in body:
            try:
                tm = int(body.get('tier_min') or 0)
                kw['tier_min'] = tm if 0 <= tm <= 10000 else 0
            except (TypeError, ValueError):
                kw['tier_min'] = 0
        if 'tier_discount' in body:
            try:
                td = float(body.get('tier_discount') or 0)
                kw['tier_discount'] = td if math.isfinite(td) and 0 <= td <= 1000000 else 0.0
            except (TypeError, ValueError):
                kw['tier_discount'] = 0.0
        db.update_delivery(pid, **kw)
        d = db.get_delivery(pid)
        host = self.headers.get('Host', '127.0.0.1:%d' % self.server.server_port)
        self._json({'ok': True, 'link': _delivery_link(d, host)})

    def _delivery_resetcode_post(self):
        body = self._read_json_body()
        pid = body.get('p') or self._query('p') or ''
        if not pid:
            self._json({'error': '缺少 p'}, 400)
            return
        db.ensure_delivery(pid)
        db.reset_delivery_code(pid)
        d = db.get_delivery(pid)
        host = self.headers.get('Host', '127.0.0.1:%d' % self.server.server_port)
        self._json({'ok': True, 'code': d['code'], 'link': _delivery_link(d, host)})

    def _delivery_confirm_post(self):
        body = self._read_json_body()
        o = db.get_order(body.get('order_id')) if body.get('order_id') else None
        if not o:
            self._json({'error': '订单不存在'}, 404)
            return
        if body.get('agree'):
            db.confirm_order(o['id'])
        elif body.get('reject'):
            db.delete_order(o['id'])
        self._json({'ok': True, 'order': _order_public(db.get_order(o['id']))})

    def _delivery_uploadqr_post(self):
        import base64
        body = self._read_json_body(max_bytes=8 * 1024 * 1024)
        pid = body.get('p') or self._query('p') or ''
        data_url = body.get('data_url') or ''
        if not pid or not data_url.startswith('data:image/'):
            self._json({'error': '需要 p 和 base64 图片'}, 400)
            return
        m = re.match(r'^data:image/(\w+);base64,(.*)$', data_url, re.S)
        if not m:
            self._json({'error': '图片格式不支持'}, 400)
            return
        ext = 'png' if m.group(1).lower() == 'png' else 'jpg'
        try:
            raw = base64.b64decode(m.group(2))
        except Exception:
            self._json({'error': '图片解码失败'}, 400)
            return
        if len(raw) > 5 * 1024 * 1024:
            self._json({'error': '图片不能超过 5 MB'}, 400)
            return
        os.makedirs(QR_DIR, exist_ok=True)
        path = os.path.join(QR_DIR, sanitize(pid) + '.' + ext)
        with open(path, 'wb') as f:
            f.write(raw)
        db.ensure_delivery(pid)
        db.update_delivery(pid, pay_qr_path=path)
        self._json({'ok': True, 'has_pay_qr': True})

    def _delivery_stats_get(self):
        pid = self._query('p') or ''
        if not pid:
            self._json({'error': '缺少 p'}, 400)
            return
        d = db.ensure_delivery(pid)
        host = self.headers.get('Host', '127.0.0.1:%d' % self.server.server_port)
        orders = []
        for o in db.list_orders(pid):
            selmap = db.get_selections(pid, o['session'])
            phs = []
            try:
                ids = json.loads(o['photo_ids'] or '[]')
            except ValueError:
                ids = []
            for phid in ids:
                st = selmap.get(phid, {})
                phs.append({'photo_id': phid, 'selected': st.get('selected', 0),
                            'note': st.get('note', '')})
            orders.append({
                'id': o['id'], 'count': o['count'],
                'free_count': int(o.get('free_count') or 0),
                'paid_count': max(0, o['count'] - int(o.get('free_count') or 0)),
                'price': o['price'], 'total': o['total'],
                'paid_amount': o['paid_amount'], 'status': o['status'],
                'customer_name': o['customer_name'], 'customer_msg': o['customer_msg'],
                'created_at': o['created_at'], 'confirmed_at': o['confirmed_at'],
                'photos': phs,
            })
        self._json({
            'config': {
                'project_id': d['project_id'], 'code': d['code'], 'title': d['title'],
                'price': d['price'] or 0.0,
                'free_count': int(d.get('free_count') or 0),
                'tier_min': int(d.get('tier_min') or 0),
                'tier_discount': float(d.get('tier_discount') or 0.0),
                'enabled': d['enabled'],
                'public_base': d['public_base'] or '',
                'has_pay_qr': bool(d['pay_qr_path']),
            },
            'link': _delivery_link(d, host),
            'orders': orders,
        })

    def _prewarm_page_get(self):
        body = PREWARM_HTML.replace('__TOKEN__', self._query('t') or '').encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.send_header('Cache-Control', 'no-store')
        self.end_headers()
        self.wfile.write(body)

    def _delivery_qr_get(self):
        pid = self._query('p') or ''
        if not pid:
            self._json({'error': '缺少 p'}, 400)
            return
        d = db.ensure_delivery(pid)
        host = self.headers.get('Host', '127.0.0.1:%d' % self.server.server_port)
        link = _delivery_link(d, host)
        try:
            import segno
        except ImportError:
            self._json({'error': '二维码库未安装: pip install segno'}, 500)
            return
        buf = io.BytesIO()
        segno.make(link, error='m').save(buf, kind='png', scale=6, dark='#1a1a1a', light=None)
        buf.seek(0)
        self._send_bytes(buf.getvalue(), 'image/png', 'delivery_qr.png')

    def _send_bytes(self, data, ctype, disposition):
        self.send_response(200)
        self.send_header('Content-Type', ctype)
        self.send_header('Content-Length', str(len(data)))
        self.send_header('Cache-Control', 'no-store')
        self.send_header('Content-Disposition',
                         "attachment; filename=\"download.bin\"; filename*=UTF-8''%s" % quote(disposition))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, fmt, *args):
        try:
            sys.stderr.write('[%s] %s\n' % (self.log_date_time_string(), fmt % args))
        except Exception:
            pass


class GalleryServer(ThreadingHTTPServer):
    daemon_threads = True
    allow_reuse_address = True

    def __init__(self, addr, ws_root, token, wm_workers=2, license_server=None):
        self._request_slots = threading.BoundedSemaphore(32)
        super().__init__(addr, Handler)
        self.ws_root = ws_root
        self.token = token
        self.license_server = license_server
        self._act_attempts = {}
        self.session_secret = _load_session_secret()
        self.image_service = ImageService(CLEAN_CACHE_DIR, workers=wm_workers)
        self.export_manager = ExportManager(self.image_service, DATA_DIR)
        from prewarm import TenantWarmer
        self.tenant_warmer = TenantWarmer(TENANT_ROOT, self.image_service, batch=2)

    def process_request(self, request, client_address):
        if not self._request_slots.acquire(blocking=False):
            self.shutdown_request(request)
            return
        try:
            super().process_request(request, client_address)
        except Exception:
            self._request_slots.release()
            raise

    def process_request_thread(self, request, client_address):
        try:
            super().process_request_thread(request, client_address)
        finally:
            self._request_slots.release()


# ---------------------------------------------------------------- 启动
def main():
    ap = argparse.ArgumentParser(description='像素蛋糕选片系统')
    ap.add_argument('--port', type=int, default=8888)
    ap.add_argument('--host', default='0.0.0.0')
    ap.add_argument('--ws', default=scanner.DEFAULT_WS)
    ap.add_argument('--token', default=None)
    ap.add_argument('--no-auth', action='store_true')
    ap.add_argument('--wm-workers', type=int, choices=(1, 2), default=2,
                    help='去水印并发线程数 (默认 2)')
    ap.add_argument('--license-server', default=None,
                    help='授权服务器地址 (软件卖给摄影师时使用, 未激活拦截管理端)')
    ap.add_argument('--data-dir', default=None,
                    help='数据目录 (默认 ./data; 打包后应指向用户目录, 如 ~/.pixcake-admin)')
    args = ap.parse_args()

    if args.data_dir:
        global DATA_DIR, CLEAN_CACHE_DIR, QR_DIR, TOKEN_FILE, PLATFORM_FILE
        DATA_DIR = os.path.abspath(args.data_dir)
        CLEAN_CACHE_DIR = os.path.join(DATA_DIR, 'clean-cache')
        QR_DIR = os.path.join(DATA_DIR, 'qr')
        TOKEN_FILE = os.path.join(DATA_DIR, 'token.txt')
        PLATFORM_FILE = os.path.join(DATA_DIR, 'platform.json')

    os.makedirs(DATA_DIR, exist_ok=True)
    db.init_db()
    _load_session_secret()

    # 多租户迁移: 旧相册归属管理员
    conn = db._connect()
    try:
        conn.execute("UPDATE projects SET owner=? WHERE owner IS NULL OR owner=''",
                     (ADMIN_OWNER,))
        conn.commit()
    finally:
        conn.close()

    token = args.token
    if token is None and not args.no_auth:
        if os.path.isfile(TOKEN_FILE):
            token = open(TOKEN_FILE, 'r').read().strip()
        if not token:
            token = secrets.token_hex(16)
            with open(TOKEN_FILE, 'w') as f:
                f.write(token)

    print('扫描工作区: %s' % args.ws)
    albums = rescan_workspace(args.ws, owner=ADMIN_OWNER)
    print('发现 %d 个相册' % len(albums))

    srv = GalleryServer((args.host, args.port), args.ws, token, args.wm_workers, args.license_server)
    from prewarm import Warmer
    srv.warmer = Warmer(args.ws, srv.image_service, interval=30, batch=4)
    srv.warmer.start()
    print('')
    print('=' * 64)
    print('像素蛋糕选片系统已启动')
    print('  自动预热 : 每 30 秒扫描新相册/新精修并预生成去水印图 (状态: /api/prewarm)')
    print('  本机地址 : http://127.0.0.1:%d' % args.port)
    print('  局域网地址: http://<本机IP>:%d' % args.port)
    print('  花生壳   : 花生壳映射本机端口 %d, 用映射后的公网地址' % args.port)
    if token:
        print('  访问令牌 : %s' % token)
    print('  项目链接 :')
    for a in albums:
        n = db.get_project(a['id'])
        cnt = n['photo_count'] if n else 0
        sel = n['sel_count'] if n else 0
        if cnt:
            print('    [%3d张/已选%d] %s%s' % (cnt, sel, a['id'], '  (http://127.0.0.1:%d/#/p/%s?t=%s)' % (args.port, a['id'], token) if token else ''))
    print('=' * 64)
    try:
        srv.serve_forever()
    except KeyboardInterrupt:
        print('\n已停止')
    finally:
        srv.warmer.stop()
        srv.server_close()
        srv.export_manager.shutdown()
        srv.image_service.shutdown()


if __name__ == '__main__':
    main()
